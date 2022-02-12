from openpyxl import load_workbook

workbook = load_workbook("Book1.xlsx")
sheet = workbook.active
a = sheet["A1:A7"]

data = []

for value in sheet.iter_rows(values_only=True):
    data.append(value)


from openpyxl import Workbook

wb = Workbook()
ws = wb.active


ws['A1'].value = sheet['A1'].value
ws['B1'] = sheet['A2'].value
ws['C1'] = sheet['A3'].value
ws['D1'] = sheet['A4'].value
ws['E1'] = sheet['A5'].value
ws['F1'] = sheet['A6'].value
ws['G1'] = sheet['A7'].value

wb.save('newFile.xlsx')
